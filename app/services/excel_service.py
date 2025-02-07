import os
import openpyxl
from app.core.database import database
from sentence_transformers import SentenceTransformer
import numpy as np

# Charger le modèle de vectorisation une seule fois pour optimiser les performances
model = SentenceTransformer("all-MiniLM-L6-v2")

def load_excel_with_hidden_content(file_name):
    """
    Charge un fichier Excel en incluant les contenus masqués.
    
    Args:
        file_name (str): Nom du fichier Excel (ex. : "example.xlsm").

    Returns:
        dict: Dictionnaire contenant les feuilles et leurs données (visibles et masquées).
    """
    try:
        # Construire le chemin complet
        base_path = os.path.join(os.getcwd(), "app", "src")
        file_path = os.path.join(base_path, file_name)

        # Charger le fichier Excel avec openpyxl
        workbook = openpyxl.load_workbook(file_path, data_only=True, keep_links=True)

        data = {}
        for sheet_name in workbook.sheetnames:
            sheet = workbook[sheet_name]
            sheet_data = []
            for row in sheet.iter_rows(values_only=True):
                sheet_data.append(row)
            data[sheet_name] = sheet_data

        return data
    except Exception as e:
        raise ValueError(f"Erreur lors du chargement du fichier Excel : {e}")

def clean_data(data):
    """
    Nettoie les données extraites, supprime les lignes vides et normalise.

    Args:
        data (list): Liste de lignes extraites d'une feuille Excel.

    Returns:
        list: Données nettoyées.
    """
    cleaned_data = []
    for row in data:
        if any(row):  # Conserver les lignes qui ne sont pas entièrement vides
            cleaned_data.append([str(cell).strip() if cell else "" for cell in row])
    return cleaned_data

def detect_macros(file_name):
    """
    Vérifie si un fichier Excel contient des macros.

    Args:
        file_name (str): Nom du fichier Excel.

    Returns:
        bool: True si des macros sont détectées, False sinon.
    """
    try:
        # Charger le fichier Excel
        base_path = os.path.join(os.getcwd(), "app", "src")
        file_path = os.path.join(base_path, file_name)
        workbook = openpyxl.load_workbook(file_path, keep_vba=True)

        return workbook.vba_archive is not None
    except Exception:
        return False

def vectorize_page(page_data):
    """
    Vectorise les données d'une page (chunk).

    Args:
        page_data (list): Données de la page.

    Returns:
        list: Embedding vectorisé pour la page.
    """
    text = " ".join([" ".join(map(str, row)) for row in page_data])
    embedding = model.encode(text)
    return embedding.tolist()  # Convertir en liste Python


def process_excel_file(file_name):
    """
    Traite un fichier Excel et génère un chunk par page (feuille).

    Args:
        file_name (str): Nom du fichier Excel.

    Returns:
        dict: Données traitées avec les embeddings.
    """
    try:
        # Charger tout le contenu, y compris les données masquées
        data = load_excel_with_hidden_content(file_name)

        processed_data = []
        for sheet_name, sheet_data in data.items():
            # Nettoyer les données
            cleaned_data = clean_data(sheet_data)
            
            # Vectoriser les données de la feuille entière
            embedding = vectorize_page(cleaned_data)

            # Ajouter les résultats
            processed_data.append({
                "sheet_name": sheet_name,
                "data": cleaned_data,
                "embedding": embedding
            })

        # Vérifier si le fichier contient des macros
        contains_macros = detect_macros(file_name)

        return {
            "status": "success",
            "processed_data": processed_data,
            "contains_macros": contains_macros
        }
    except Exception as e:
        return {"status": "error", "message": str(e)}

async def save_pages_to_db(pages, collection_name="test-chunks"):
    """
    Sauvegarde les pages (chunks) dans une collection MongoDB.

    Args:
        pages (list): Liste des pages (feuilles) à sauvegarder.
        collection_name (str): Nom de la collection MongoDB.
    """
    try:
        collection = database[collection_name]
        await collection.insert_many(pages)
        return {"status": "success", "message": f"{len(pages)} pages insérées dans {collection_name}."}
    except Exception as e:
        return {"status": "error", "message": str(e)}

async def save_chunks_to_db(chunks, collection_name="test-chunks"):
    """
    Sauvegarde les chunks dans une collection MongoDB.

    Args:
        chunks (list): Liste des chunks à sauvegarder.
        collection_name (str): Nom de la collection MongoDB.
    """
    try:
        for chunk in chunks:
            if isinstance(chunk.get("embedding"), np.ndarray):
                chunk["embedding"] = chunk["embedding"].tolist()  # Convertir si nécessaire

        collection = database[collection_name]
        await collection.insert_many(chunks)
        return {"status": "success", "message": f"{len(chunks)} chunks insérés dans {collection_name}."}
    except Exception as e:
        return {"status": "error", "message": str(e)}
